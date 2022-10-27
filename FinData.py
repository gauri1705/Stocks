from openpyxl import load_workbook
from pathlib import Path
import yfinance as yf

class FinancialData():
    def __init__(self):
        self.xlsx = load_workbook(Path(__file__).parent / 'Index  Holdings - Daily Performance 2.xlsx')
        self.sheet = self.xlsx['Yahoo_Data']
        self.tickers = []
        #self.tickers = [c.value for c in self.sheet.columns[0]]
        for col in self.sheet['B'][1:]:
            self.tickers.append(col.value)

    def assign_cell_value(self, sheet, row, col, val):
        if val is not None:
            sheet.cell(row=row, column=col).value = val

    def load_ticker(self, row: int, ticker: str):
        t_par = yf.Ticker(ticker)
        t_info = t_par.info
        print(f"ticker: {ticker}")
        print(f"t_info:", t_info)
        print(f"t_info 52 week high:", t_info.get('fiftyTwoWeekHigh'))
        print(f"t_info 52 week low:", t_info.get('fiftyTwoWeekLow'))
        current_price = round(t_par.info.get('regularMarketPrice'), 2)
        print(f'current_price: {current_price}')
        self.assign_cell_value(self.sheet, row, 3, current_price)
        previous_close = round(t_par.info.get('regularMarketPreviousClose'), 2)
        print(f'previous_close: {previous_close}')
        self.assign_cell_value(self.sheet, row, 4, previous_close)
        up_down = round((current_price - previous_close), 2)
        print(f'up down : {up_down}')
        self.assign_cell_value(self.sheet, row, 6, up_down)
        up_down_per = (up_down/previous_close)*100
        self.assign_cell_value(self.sheet, row, 5, up_down_per)
        self.assign_cell_value(self.sheet, row, 7, t_info.get('regularMarketVolume'))
        self.assign_cell_value(self.sheet, row, 8, round(t_info.get('fiftyTwoWeekHigh'), 2))
        self.assign_cell_value(self.sheet, row, 9, round(t_info.get('fiftyTwoWeekLow'), 2))

    def load_all_tickers(self):
        for row, ticker in enumerate(self.tickers, start=2):
            self.load_ticker(row, ticker)

    def save_data(self, name_xls):
        self.xlsx.save(Path(__file__).parent / name_xls)